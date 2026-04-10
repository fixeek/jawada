"""
Jawda KPI Platform — FastAPI Backend
"""
import logging
import sys

# Configure structured logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(name)s: %(message)s',
    datefmt='%Y-%m-%dT%H:%M:%S',
    handlers=[logging.StreamHandler(sys.stdout)],
)
log = logging.getLogger("jawda")

from fastapi import FastAPI, UploadFile, File, HTTPException, Form, Depends, Header, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, FileResponse, Response
import io
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
import pandas as pd
import tempfile, os, json, uuid, secrets, string
from datetime import datetime
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from engine.kpi_engine import run_all_kpis, load_file, normalise_df, map_columns
from typing import Optional

# Database — try to connect, fall back to in-memory if unavailable
USE_DB = False
try:
    from database import (init_db, db_cursor, save_results, get_facility_history, get_facility_history_by_id,
                          save_facility_col_mapping, update_submission_status,
                          log_audit, get_audit_log,
                          get_platform_stats, get_platform_audit_log, get_system_health,
                          create_notification, get_notifications, mark_notifications_read,
                          get_unread_count, generate_deadline_notifications,
                          create_invitation, get_invitation, use_invitation,
                          get_all_clinics_kpi_summary)
    init_db()
    USE_DB = True
    log.info("✓ Connected to PostgreSQL")
except Exception as e:
    log.warning(f"⚠ Database unavailable, using in-memory: {e}")
    def log_audit(*args, **kwargs): pass
    def get_audit_log(*args, **kwargs): return []
    def get_platform_stats(*args, **kwargs): return {}
    def get_platform_audit_log(*args, **kwargs): return []
    def get_system_health(*args, **kwargs): return {}

# Auth module
try:
    from auth import (
        verify_password, create_token, decode_token, get_user_by_email, get_user_by_id,
        create_user, list_users_for_facility, list_all_facilities, create_facility,
        update_user_password, update_last_login, delete_user, deactivate_user,
        reactivate_user, deactivate_facility, reactivate_facility, seed_super_admin,
        EmailAlreadyExistsError,
        ROLE_SUPER_ADMIN, ROLE_CLINIC_ADMIN, ROLE_QUALITY_OFFICER, ROLE_VIEWER, ALL_ROLES
    )
    USE_AUTH = USE_DB
    if USE_AUTH:
        seed_super_admin()
        log.info("✓ Auth module ready")
except Exception as e:
    log.warning(f"⚠ Auth unavailable: {e}")
    USE_AUTH = False

# Email service (graceful fallback if not configured)
try:
    from email_service import (
        send_welcome_email, send_kpi_calculation_email, send_password_changed_email,
        send_password_reset_email,
        is_enabled as email_is_enabled
    )
    USE_EMAIL = True
except Exception as e:
    log.warning(f"⚠ Email service unavailable: {e}")
    USE_EMAIL = False
    def send_welcome_email(*args, **kwargs): return False
    def send_kpi_calculation_email(*args, **kwargs): return False
    def send_password_changed_email(*args, **kwargs): return False
    def send_password_reset_email(*args, **kwargs): return False

app = FastAPI(
    title="Jawda KPI Platform",
    description="DOH Jawda KPI Reporting Engine — v1.4 (Q1 2025)",
    version="1.0.0"
)


# ─────────────────────────────────────────────────────────────────────────
# Global exception handlers — always return clean {"detail": "..."} JSON
# ─────────────────────────────────────────────────────────────────────────

from fastapi.responses import JSONResponse as _JSONResponse
from fastapi.requests import Request as _Request
from fastapi.exceptions import RequestValidationError as _RequestValidationError
from starlette.exceptions import HTTPException as _StarletteHTTPException
import traceback as _traceback


@app.exception_handler(_StarletteHTTPException)
async def _http_exception_handler(request: _Request, exc: _StarletteHTTPException):
    """Format HTTPExceptions as clean {"detail": "..."}."""
    return _JSONResponse(
        status_code=exc.status_code,
        content={"detail": exc.detail if isinstance(exc.detail, str) else str(exc.detail)},
    )


@app.exception_handler(_RequestValidationError)
async def _validation_exception_handler(request: _Request, exc: _RequestValidationError):
    """Convert pydantic validation errors to a single readable message."""
    errors = exc.errors()
    if errors:
        first = errors[0]
        loc = ".".join(str(p) for p in first.get("loc", []) if p != "body")
        msg = first.get("msg", "Validation failed")
        if loc:
            detail = f"{loc}: {msg}"
        else:
            detail = msg
    else:
        detail = "Validation failed"
    return _JSONResponse(status_code=422, content={"detail": detail})


@app.exception_handler(Exception)
async def _unhandled_exception_handler(request: _Request, exc: Exception):
    """Catch-all: log full traceback, return clean message to client."""
    log.error(f"Unhandled error on {request.method} {request.url.path}: {exc}")
    log.error(_traceback.format_exc())

    # psycopg2 unique violation
    err_str = str(exc).lower()
    if "duplicate key" in err_str:
        if "email" in err_str:
            return _JSONResponse(status_code=409, content={"detail": "A user with this email already exists."})
        if "name_lower" in err_str:
            return _JSONResponse(status_code=409, content={"detail": "A facility with this name already exists."})
        return _JSONResponse(status_code=409, content={"detail": "This record already exists."})

    if "violates foreign key" in err_str:
        return _JSONResponse(status_code=400, content={"detail": "Referenced record does not exist."})

    if "violates not-null" in err_str or "null value in column" in err_str:
        return _JSONResponse(status_code=400, content={"detail": "Required field is missing."})

    if "connection refused" in err_str or "could not connect" in err_str or "connection timed out" in err_str:
        return _JSONResponse(status_code=503, content={"detail": "Database is temporarily unavailable. Please try again."})

    # Generic fallback — don't leak internals to client
    return _JSONResponse(
        status_code=500,
        content={"detail": "An unexpected error occurred. Please try again or contact support if the issue persists."},
    )

# CORS — locked to production domain + localhost for dev
ALLOWED_ORIGINS_ENV = os.environ.get("ALLOWED_ORIGINS", "")
if ALLOWED_ORIGINS_ENV:
    allowed_origins = [o.strip() for o in ALLOWED_ORIGINS_ENV.split(",") if o.strip()]
else:
    allowed_origins = [
        "https://jawda.trizodiac.com",
        "http://localhost:5173",
        "http://localhost:3000",
    ]

app.add_middleware(
    CORSMiddleware,
    allow_origins=allowed_origins,
    allow_credentials=True,
    allow_methods=["GET", "POST", "PUT", "DELETE", "PATCH", "OPTIONS"],
    allow_headers=["Authorization", "Content-Type", "Accept"],
)

# In-memory store for demo (replace with PostgreSQL in production)
results_store = {}         # run_id -> results
facility_history = {}      # facility_name_lower -> { quarter -> results }
# Temp store for validated files awaiting processing
validated_files = {}
SESSION_MAX_AGE_MINUTES = 30


def _cleanup_stale_sessions():
    """Remove validated file sessions older than SESSION_MAX_AGE_MINUTES.
    Called before each new validate to prevent unbounded growth."""
    now = datetime.now()
    stale = []
    for sid, session in validated_files.items():
        try:
            created = datetime.fromisoformat(session.get("created", ""))
            if (now - created).total_seconds() > SESSION_MAX_AGE_MINUTES * 60:
                stale.append(sid)
        except:
            stale.append(sid)
    for sid in stale:
        session = validated_files.pop(sid, {})
        for p in list(session.get("paths", {}).values()) + list(session.get("prev_paths", {}).values()):
            try:
                os.unlink(p)
            except:
                pass
    if stale:
        log.info(f"Cleaned up {len(stale)} stale upload sessions")


def _save_to_history(facility: str, quarter: str, results: dict):
    """Store results to database (if available) and in-memory for trend comparison."""
    if USE_DB:
        try:
            save_results(facility, results)
            # Auto-save column mapping as clinic's default (first upload only)
            col_mapping = results.get("col_mapping")
            if col_mapping:
                from database import get_or_create_facility
                fid = get_or_create_facility(facility)
                save_facility_col_mapping(fid, col_mapping)
        except Exception as e:
            log.error(f"DB save error: {e}")

    # Also keep in-memory for current session
    key = facility.strip().lower()
    if key not in facility_history:
        facility_history[key] = {}
    facility_history[key][quarter] = {
        "quarter": quarter,
        "saved_at": datetime.now().isoformat(),
        "jawda_summary": results.get("jawda_summary", {}),
        "kpis": {
            kpi_id: {
                "numerator": kpi.get("numerator", 0),
                "denominator": kpi.get("denominator", 0),
                "percentage": kpi.get("percentage", 0),
                "status": kpi.get("status", ""),
                "meets_target": kpi.get("meets_target"),
                "target": kpi.get("target", 0),
                "target_direction": kpi.get("target_direction", "higher"),
                "title": kpi.get("title", ""),
                "domain": kpi.get("domain", ""),
            }
            for kpi_id, kpi in results.get("kpis", {}).items()
            if not kpi_id.startswith("ERROR")
        },
    }


def _safe_to_datetime(val):
    """Convert a value to datetime, handling both already-parsed Timestamps and strings.
    Uses dayfirst=True for strings (UAE standard DD/MM/YYYY)."""
    if isinstance(val, pd.Timestamp):
        return val
    if isinstance(val, datetime):
        return pd.Timestamp(val)
    try:
        return pd.to_datetime(val, dayfirst=True)
    except:
        return None


def _detect_quarters(df: pd.DataFrame, date_col: str) -> list:
    """Detect which quarters are present in a date column."""
    quarters = {}
    for val in df[date_col].dropna():
        dt = _safe_to_datetime(val)
        if dt is None:
            continue
        q_label = f"Q{(dt.month - 1) // 3 + 1} {dt.year}"
        quarters[q_label] = quarters.get(q_label, 0) + 1
    # Sort by year then quarter
    sorted_q = sorted(quarters.items(), key=lambda x: (int(x[0].split()[1]), x[0].split()[0]), reverse=True)
    return [{"quarter": q, "record_count": c} for q, c in sorted_q]


def _filter_by_quarter(df: pd.DataFrame, date_col: str, quarter: str) -> pd.DataFrame:
    """Filter dataframe to only rows matching the given quarter (e.g. 'Q1 2025')."""
    parts = quarter.split()
    if len(parts) != 2:
        return df
    q_num, year = int(parts[0][1]), int(parts[1])
    month_start = (q_num - 1) * 3 + 1
    month_end = month_start + 2

    mask = []
    for val in df[date_col]:
        dt = _safe_to_datetime(val)
        if dt is None:
            mask.append(False)
        else:
            mask.append(dt.year == year and month_start <= dt.month <= month_end)
    return df[mask].reset_index(drop=True)


@app.get("/api/template/download")
def download_template():
    """Download a blank Excel template showing expected column headers."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from engine.kpi_engine import FIELD_DEFINITIONS

    wb = Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="0D2137", end_color="0D2137", fill_type="solid")
    desc_font = Font(italic=True, color="888888", size=9)

    sheets_config = {
        "KPI Data": ["patient_id", "file_no", "date_of_service", "date_of_birth", "age", "gender",
                      "bmi", "bp_reading", "hba1c", "diabetes_flag", "htn_flag", "egfr", "uacr",
                      "opioid", "pregnancy", "management_plan"],
        "Visit Details": ["patient_id", "date_of_service", "icd_code", "icd_secondary",
                          "cpt_code", "drug_name", "drug_class", "days_supplied", "pregnancy"],
        "Time Data": ["file_no", "date_of_service", "registration_ts", "consult_ts", "wait_minutes"],
        "E-Claims": ["patient_id", "date_of_service", "age", "date_of_birth", "gender", "visit_type"],
    }

    examples = {
        "patient_id": "12345", "file_no": "67890", "date_of_service": "15/07/2025",
        "date_of_birth": "20/05/1980", "age": "45", "gender": "F",
        "bmi": "27.5", "bp_reading": "130/80", "hba1c": "7.2",
        "diabetes_flag": "YES", "htn_flag": "NO", "egfr": "85",
        "uacr": "25", "opioid": "NO", "pregnancy": "NO",
        "management_plan": "Documented", "icd_code": "E11.9",
        "icd_secondary": "I10", "cpt_code": "99213", "drug_name": "Metformin",
        "drug_class": "Antidiabetic", "days_supplied": "30",
        "registration_ts": "15/07/2025 08:30", "consult_ts": "15/07/2025 09:15",
        "wait_minutes": "45", "visit_type": "O",
    }

    for sheet_name, fields in sheets_config.items():
        ws = wb.create_sheet(sheet_name)
        for col_idx, field_key in enumerate(fields, 1):
            fd = FIELD_DEFINITIONS.get(field_key, {})
            # Row 1: column header (preferred alias)
            cell = ws.cell(row=1, column=col_idx, value=fd.get("aliases", [field_key])[0] if fd.get("aliases") else field_key)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            # Row 2: description
            ws.cell(row=2, column=col_idx, value=fd.get("label", field_key)).font = desc_font
            # Row 3: required/optional
            ws.cell(row=3, column=col_idx, value="REQUIRED" if fd.get("required") else "Optional").font = desc_font
            # Row 4: affected KPIs
            ws.cell(row=4, column=col_idx, value=", ".join(fd.get("kpis", []))).font = desc_font
            # Row 5: example value
            ws.cell(row=5, column=col_idx, value=examples.get(field_key, ""))
            # Column width
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="Jawda-KPI-Data-Template.xlsx"'},
    )


@app.get("/api")
def root():
    return {"status": "ok", "product": "Jawda KPI Platform", "developer": "TriZodiac", "version": "1.0.0"}

@app.get("/health")
def health():
    return {"status": "healthy", "timestamp": datetime.now().isoformat()}


# ─────────────────────────────────────────────────────────────────────────
# AUTHENTICATION
# ─────────────────────────────────────────────────────────────────────────

class LoginRequest(BaseModel):
    email: str
    password: str

class CreateFacilityRequest(BaseModel):
    name: str
    license_no: Optional[str] = None
    doh_facility_id: Optional[str] = None
    address: Optional[str] = None
    contact_name: Optional[str] = None
    contact_email: Optional[str] = None
    contact_phone: Optional[str] = None
    admin_email: str
    admin_password: str
    admin_full_name: str

class CreateUserRequest(BaseModel):
    email: str
    password: str
    full_name: str
    role: str
    facility_id: Optional[int] = None

class ChangePasswordRequest(BaseModel):
    new_password: str

class ResetPasswordRequest(BaseModel):
    new_password: Optional[str] = None  # if omitted, server generates one

class UpdateSubmissionRequest(BaseModel):
    quarter: str
    status: str  # calculated, under_review, approved, submitted, accepted
    notes: Optional[str] = ""

class CreateInvitationRequest(BaseModel):
    facility_name: str
    admin_email: str
    admin_full_name: Optional[str] = ""
    license_no: Optional[str] = None
    doh_facility_id: Optional[str] = None

class AcceptInvitationRequest(BaseModel):
    password: str


def get_current_user(authorization: Optional[str] = Header(None)) -> dict:
    """Dependency: extract and verify JWT token, return user dict."""
    if not USE_AUTH:
        # Auth disabled — return a fake super admin (POC backwards compat)
        return {"id": 0, "email": "anonymous", "role": ROLE_SUPER_ADMIN, "facility_id": None}

    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(401, "Authentication required")

    token = authorization[7:]
    payload = decode_token(token)
    if not payload:
        raise HTTPException(401, "Invalid or expired token")

    user = get_user_by_id(int(payload["sub"]))
    if not user:
        raise HTTPException(401, "User not found")
    if not user.get("is_active", True):
        raise HTTPException(403, "User account is disabled")
    return user


def require_role(*roles):
    """Dependency factory: require user to have one of the given roles."""
    def checker(user: dict = Depends(get_current_user)) -> dict:
        if user["role"] not in roles:
            raise HTTPException(403, f"Requires one of: {', '.join(roles)}")
        return user
    return checker


def require_super_admin(user: dict = Depends(get_current_user)) -> dict:
    if user["role"] != ROLE_SUPER_ADMIN:
        raise HTTPException(403, "Super admin only")
    return user


# In-memory rate limiter for login (per IP)
from collections import defaultdict
_login_attempts = defaultdict(list)  # ip -> [timestamps]
LOGIN_MAX_ATTEMPTS = 5
LOGIN_WINDOW_SECONDS = 300  # 5 minutes


def _check_login_rate_limit(ip: str):
    """Raise 429 if too many login attempts from this IP."""
    now = datetime.utcnow().timestamp()
    cutoff = now - LOGIN_WINDOW_SECONDS
    attempts = [t for t in _login_attempts[ip] if t > cutoff]
    _login_attempts[ip] = attempts
    if len(attempts) >= LOGIN_MAX_ATTEMPTS:
        raise HTTPException(429, "Too many login attempts. Please try again in a few minutes.")


def _record_login_attempt(ip: str):
    _login_attempts[ip].append(datetime.utcnow().timestamp())


def _get_client_ip(request: Request) -> str:
    """Extract client IP, respecting X-Forwarded-For from Container Apps."""
    fwd = request.headers.get("x-forwarded-for")
    if fwd:
        return fwd.split(",")[0].strip()
    return request.client.host if request.client else "unknown"


@app.post("/api/auth/login")
def login(req: LoginRequest, request: Request):
    """Login with email + password. Returns JWT token."""
    if not USE_AUTH:
        raise HTTPException(503, "Authentication system unavailable")

    client_ip = _get_client_ip(request)
    user_agent = request.headers.get("user-agent", "")
    _check_login_rate_limit(client_ip)
    _record_login_attempt(client_ip)

    user = get_user_by_email(req.email)
    if not user or not verify_password(req.password, user["password_hash"]):
        # Audit failed login attempts
        try:
            log_audit(
                facility_name="",
                action="login_failed",
                details={"email": req.email[:120]},
                user_email=req.email,
                ip_address=client_ip,
                user_agent=user_agent,
            )
        except:
            pass
        raise HTTPException(401, "Invalid email or password")
    if not user.get("is_active", True):
        try:
            log_audit(
                facility_name=user.get("facility_name") or "",
                action="login_blocked",
                details={"reason": "account_disabled"},
                user_email=user["email"],
                ip_address=client_ip,
                user_agent=user_agent,
            )
        except:
            pass
        raise HTTPException(403, "Account disabled")

    update_last_login(user["id"])

    # Audit successful login
    try:
        log_audit(
            facility_name=user.get("facility_name") or "",
            action="login_success",
            details={"role": user["role"]},
            user_email=user["email"],
            ip_address=client_ip,
            user_agent=user_agent,
        )
    except:
        pass

    token = create_token(
        user_id=user["id"],
        email=user["email"],
        role=user["role"],
        facility_id=user.get("facility_id"),
    )

    return {
        "token": token,
        "user": {
            "id": user["id"],
            "email": user["email"],
            "full_name": user.get("full_name"),
            "role": user["role"],
            "facility_id": user.get("facility_id"),
            "facility_name": user.get("facility_name"),
            "must_change_password": user.get("must_change_password", False),
        }
    }


@app.get("/api/auth/me")
def get_me(user: dict = Depends(get_current_user)):
    """Get current authenticated user info."""
    return user


@app.post("/api/auth/change-password")
def change_password(req: ChangePasswordRequest, user: dict = Depends(get_current_user)):
    """Change own password."""
    if not USE_AUTH:
        raise HTTPException(503, "Auth unavailable")
    if len(req.new_password) < 8:
        raise HTTPException(400, "Password must be at least 8 characters")
    update_user_password(user["id"], req.new_password)

    try:
        send_password_changed_email(
            to_email=user["email"],
            full_name=user.get("full_name") or user["email"],
        )
    except Exception as e:
        log.error(f"Failed to send password change email: {e}")

    return {"success": True}


# ─────────────────────────────────────────────────────────────────────────
# CLINIC / FACILITY MANAGEMENT (super admin)
# ─────────────────────────────────────────────────────────────────────────

@app.get("/api/admin/stats")
def admin_platform_stats(user: dict = Depends(require_super_admin)):
    """Super admin: platform-wide statistics."""
    return get_platform_stats()


@app.get("/api/admin/clinics-kpi")
def admin_clinics_kpi(user: dict = Depends(require_super_admin)):
    """Super admin: KPI summary per clinic."""
    try:
        return {"clinics": get_all_clinics_kpi_summary()}
    except Exception as e:
        log.error(f"Clinics KPI summary failed: {e}")
        return {"clinics": []}


# ── Invitations ─────────────────────────────────────────────────────

@app.post("/api/admin/invitations")
def send_invitation(req: CreateInvitationRequest, user: dict = Depends(require_super_admin)):
    """Super admin sends a clinic onboarding invitation."""
    token = create_invitation(
        facility_name=req.facility_name,
        admin_email=req.admin_email,
        admin_full_name=req.admin_full_name or "",
        license_no=req.license_no,
        doh_facility_id=req.doh_facility_id,
        created_by=user.get("id"),
    )
    invite_url = f"{os.environ.get('APP_URL', 'https://jawda.trizodiac.com')}?invite={token}"

    # Send invitation email
    try:
        from email_service import _send, _wrap
        body = f'''
        <p>Hello <strong>{req.admin_full_name or req.admin_email}</strong>,</p>
        <p>You have been invited to join the <strong>Jawda KPI Platform</strong> as the
        Clinic Admin for <strong>{req.facility_name}</strong>.</p>
        <p>Click the button below to set your password and activate your account:</p>
        '''
        _send(
            to_email=req.admin_email,
            subject=f"You're invited to Jawda KPI Platform — {req.facility_name}",
            html=_wrap("Clinic Invitation", body, cta_text="Accept Invitation", cta_url=invite_url),
            plain_text=f"You're invited to Jawda KPI as admin for {req.facility_name}. Accept at {invite_url}",
        )
    except Exception as e:
        log.error(f"Invitation email failed: {e}")

    return {"success": True, "token": token, "invite_url": invite_url}


@app.get("/api/invite/{token}")
def get_invite(token: str):
    """Public: get invitation details by token."""
    invite = get_invitation(token)
    if not invite:
        raise HTTPException(404, "Invitation not found, expired, or already used.")
    return {
        "facility_name": invite["facility_name"],
        "admin_email": invite["admin_email"],
        "admin_full_name": invite.get("admin_full_name", ""),
        "expires_at": invite["expires_at"],
    }


@app.post("/api/invite/{token}/accept")
def accept_invite(token: str, req: AcceptInvitationRequest):
    """Public: accept invitation — creates facility + admin user, returns JWT."""
    invite = get_invitation(token)
    if not invite:
        raise HTTPException(404, "Invitation not found, expired, or already used.")

    if len(req.password) < 8:
        raise HTTPException(400, "Password must be at least 8 characters")

    # Create facility
    facility_id = create_facility(
        name=invite["facility_name"],
        license_no=invite.get("license_no"),
        doh_facility_id=invite.get("doh_facility_id"),
    )

    # Create admin user
    try:
        new_user = create_user(
            email=invite["admin_email"],
            password=req.password,
            full_name=invite.get("admin_full_name", ""),
            role=ROLE_CLINIC_ADMIN,
            facility_id=facility_id,
            must_change_password=False,  # They just set it
        )
    except EmailAlreadyExistsError:
        raise HTTPException(409, "This email is already registered.")

    # Mark invitation as used
    use_invitation(token)

    # Generate JWT token so user is logged in immediately
    jwt_token = create_token(new_user["id"])

    return {
        "success": True,
        "token": jwt_token,
        "user": new_user,
        "facility_name": invite["facility_name"],
    }


@app.get("/api/clinic/benchmark/{kpi_id}")
def clinic_benchmark(kpi_id: str, user: dict = Depends(get_current_user)):
    """Get benchmark data for a KPI — how this clinic compares to others."""
    facility_id = user.get("facility_id")
    if not facility_id or not USE_DB:
        return {"available": False}
    try:
        with db_cursor(dict_cursor=True) as (conn, cur):
            # Get all clinics' latest results for this KPI
            cur.execute("""
                SELECT DISTINCT ON (facility_id)
                    facility_id, percentage, meets_target, denominator
                FROM kpi_results
                WHERE kpi_id = %s AND denominator > 0 AND status = 'calculated'
                ORDER BY facility_id, created_at DESC
            """, (kpi_id,))
            rows = cur.fetchall()
            if len(rows) < 2:
                return {"available": False, "reason": "Need 2+ clinics with data for benchmarking"}

            all_pcts = sorted([r["percentage"] for r in rows])
            my_pct = next((r["percentage"] for r in rows if r["facility_id"] == facility_id), None)
            if my_pct is None:
                return {"available": False}

            rank = sum(1 for p in all_pcts if p > my_pct) + 1
            percentile = round((1 - (rank - 1) / len(all_pcts)) * 100)
            avg = round(sum(all_pcts) / len(all_pcts), 1)

            return {
                "available": True,
                "total_clinics": len(rows),
                "your_percentage": my_pct,
                "rank": rank,
                "percentile": percentile,
                "average": avg,
                "min": all_pcts[0],
                "max": all_pcts[-1],
            }
    except Exception as e:
        log.error(f"Benchmark query failed: {e}")
        return {"available": False}


@app.post("/api/clinic/report/email")
def email_report(quarter: str = Form(...), user: dict = Depends(get_current_user)):
    """Generate PDF report and email it to the current user."""
    from report_generator import generate_pdf
    facility_id = user.get("facility_id")
    facility_name = user.get("facility_name", "Clinic")
    user_email = user.get("email")

    if not facility_id:
        raise HTTPException(400, "No facility associated")
    if not user_email:
        raise HTTPException(400, "No email on your account")

    history = {}
    if USE_DB:
        try:
            history = get_facility_history_by_id(facility_id)
        except:
            pass
    if quarter not in history:
        raise HTTPException(404, f"No data for {quarter}")

    q_data = history[quarter]
    try:
        pdf_bytes, _ = generate_pdf(
            facility_name=facility_name, quarter=quarter,
            kpis=q_data.get("kpis", {}), summary=q_data.get("jawda_summary", {}),
            history=history, generated_by=user.get("full_name", user_email),
        )
    except Exception as e:
        raise HTTPException(500, f"PDF generation failed: {e}")

    # Send email with PDF attachment
    try:
        from email_service import _send, _wrap
        body = f'''
        <p>Hello <strong>{user.get("full_name", user_email)}</strong>,</p>
        <p>Your Jawda KPI report for <strong>{facility_name} — {quarter}</strong> is attached.</p>
        <p>Sign in to view the full interactive dashboard with trends, what-if simulator, and action plan.</p>
        '''
        # Azure Communication Services doesn't support attachments in the basic API
        # Instead, regenerate via link
        app_url = os.environ.get("APP_URL", "https://jawda.trizodiac.com")
        _send(
            to_email=user_email,
            subject=f"Jawda KPI Report — {facility_name} ({quarter})",
            html=_wrap(f"{quarter} KPI Report", body, cta_text="View Dashboard", cta_url=app_url),
            plain_text=f"Your Jawda KPI report for {facility_name} ({quarter}) is ready. View at {app_url}",
        )
        return {"success": True, "message": f"Report notification sent to {user_email}"}
    except Exception as e:
        log.error(f"Email report failed: {e}")
        raise HTTPException(500, f"Email failed: {str(e)}")


@app.get("/api/clinic/recommendations")
def clinic_recommendations(user: dict = Depends(get_current_user)):
    """Get smart recommendations and predictions for the clinic."""
    facility_id = user.get("facility_id")
    if not facility_id:
        return {"recommendations": [], "predictions": []}
    try:
        history = get_facility_history_by_id(facility_id) if USE_DB else {}
        quarters = sorted(history.keys())
        if not quarters:
            return {"recommendations": [], "predictions": []}
        latest_q = quarters[-1]
        kpis = history[latest_q].get("kpis", {})
        summary = history[latest_q].get("jawda_summary", {})
        from recommendations import generate_recommendations, generate_predictions
        recs = generate_recommendations(kpis, summary)
        preds = generate_predictions(history) if len(quarters) > 1 else []
        return {"recommendations": recs, "predictions": preds}
    except Exception as e:
        log.error(f"Recommendations endpoint failed: {e}")
        return {"recommendations": [], "predictions": []}


@app.get("/api/clinic/doctors")
def clinic_doctors(user: dict = Depends(get_current_user)):
    """Get doctor breakdown from the latest calculation."""
    facility_id = user.get("facility_id")
    if not facility_id:
        return {"available": False}
    # Get latest upload results which include doctor_breakdown
    history = {}
    if USE_DB:
        try:
            history = get_facility_history_by_id(facility_id)
        except:
            pass
    quarters = sorted(history.keys())
    if not quarters:
        return {"available": False}
    latest = history[quarters[-1]]
    # doctor_breakdown is only in fresh calculation results, not stored in DB
    # Return what we have from the latest history
    return {"available": False, "note": "Doctor breakdown is shown in dashboard after fresh calculation."}


@app.get("/api/admin/audit")
def admin_platform_audit(limit: int = 100, user: dict = Depends(require_super_admin)):
    """Super admin: platform-wide audit log across all clinics."""
    return {"entries": get_platform_audit_log(limit)}


@app.get("/api/admin/health")
def admin_system_health(user: dict = Depends(require_super_admin)):
    """Super admin: system health and database stats."""
    return get_system_health()


@app.get("/api/admin/facilities")
def list_facilities(user: dict = Depends(require_super_admin)):
    """Super admin: list all facilities."""
    return {"facilities": list_all_facilities()}


@app.post("/api/admin/facilities")
def onboard_facility(req: CreateFacilityRequest, user: dict = Depends(require_super_admin)):
    """Super admin: onboard a new clinic with its first admin user."""
    # Check if facility already exists
    facility_id = create_facility(
        name=req.name,
        license_no=req.license_no,
        doh_facility_id=req.doh_facility_id,
        address=req.address,
        contact_name=req.contact_name,
        contact_email=req.contact_email,
        contact_phone=req.contact_phone,
    )

    # Create the clinic admin user
    try:
        admin_user = create_user(
            email=req.admin_email,
            password=req.admin_password,
            full_name=req.admin_full_name,
            role=ROLE_CLINIC_ADMIN,
            facility_id=facility_id,
            must_change_password=True,
        )
    except EmailAlreadyExistsError as e:
        raise HTTPException(409, str(e))
    except ValueError as e:
        raise HTTPException(400, str(e))

    # Send welcome email to the new clinic admin
    try:
        send_welcome_email(
            to_email=req.admin_email,
            full_name=req.admin_full_name,
            temp_password=req.admin_password,
            role=ROLE_CLINIC_ADMIN,
            facility_name=req.name,
        )
    except Exception as e:
        log.error(f"Failed to send welcome email: {e}")

    return {
        "facility_id": facility_id,
        "admin_user": admin_user,
    }


# ─────────────────────────────────────────────────────────────────────────
# USER MANAGEMENT (clinic admin)
# ─────────────────────────────────────────────────────────────────────────

@app.get("/api/users")
def list_my_users(facility_id: Optional[int] = None, user: dict = Depends(get_current_user)):
    """List users.
    - Clinic admin: lists own facility users (facility_id ignored)
    - Super admin: must pass facility_id to list a specific clinic's users"""
    if user["role"] == ROLE_SUPER_ADMIN:
        if not facility_id:
            return {"users": []}
        return {"users": list_users_for_facility(facility_id)}

    if not user.get("facility_id"):
        return {"users": []}

    return {"users": list_users_for_facility(user["facility_id"])}


@app.post("/api/users")
def create_team_user(req: CreateUserRequest, user: dict = Depends(get_current_user)):
    """Clinic admin creates a user in their own facility."""
    if user["role"] not in [ROLE_SUPER_ADMIN, ROLE_CLINIC_ADMIN]:
        raise HTTPException(403, "Only admins can create users")

    # Clinic admins can only create users in their own facility
    if user["role"] == ROLE_CLINIC_ADMIN:
        if req.facility_id and req.facility_id != user.get("facility_id"):
            raise HTTPException(403, "Cannot create users in other facilities")
        facility_id = user.get("facility_id")
        # Clinic admins can't create super admins
        if req.role == ROLE_SUPER_ADMIN:
            raise HTTPException(403, "Cannot create super admin users")
    else:
        facility_id = req.facility_id

    try:
        new_user = create_user(
            email=req.email,
            password=req.password,
            full_name=req.full_name,
            role=req.role,
            facility_id=facility_id,
            must_change_password=True,
        )
    except EmailAlreadyExistsError as e:
        raise HTTPException(409, str(e))
    except ValueError as e:
        raise HTTPException(400, str(e))

    # Send welcome email with credentials
    facility_name_for_email = user.get("facility_name") if facility_id == user.get("facility_id") else None
    if not facility_name_for_email and facility_id:
        try:
            target_user = get_user_by_id(new_user["id"])
            facility_name_for_email = target_user.get("facility_name")
        except:
            pass
    try:
        send_welcome_email(
            to_email=req.email,
            full_name=req.full_name,
            temp_password=req.password,
            role=req.role,
            facility_name=facility_name_for_email,
        )
    except Exception as e:
        log.error(f"Failed to send welcome email: {e}")

    return new_user


@app.delete("/api/users/{user_id}")
def remove_user(user_id: int, user: dict = Depends(get_current_user)):
    """Delete a user (admin only)."""
    if user["role"] not in [ROLE_SUPER_ADMIN, ROLE_CLINIC_ADMIN]:
        raise HTTPException(403, "Only admins can delete users")
    if user["id"] == user_id:
        raise HTTPException(400, "Cannot delete yourself")

    target = get_user_by_id(user_id)
    if not target:
        raise HTTPException(404, "User not found")

    # Clinic admins can only delete users in their own facility
    if user["role"] == ROLE_CLINIC_ADMIN:
        if target.get("facility_id") != user.get("facility_id"):
            raise HTTPException(403, "Cannot delete users from other facilities")

    delete_user(user_id)
    return {"success": True}


def _generate_temp_password(length: int = 12) -> str:
    """Generate a readable temporary password (no ambiguous characters)."""
    alphabet = string.ascii_letters.replace("l", "").replace("I", "").replace("O", "") \
        + string.digits.replace("0", "").replace("1", "") + "!@#$%"
    return "".join(secrets.choice(alphabet) for _ in range(length))


@app.post("/api/users/{user_id}/reset-password")
def reset_user_password(user_id: int, req: ResetPasswordRequest, request: Request,
                        user: dict = Depends(get_current_user)):
    """Admin-initiated password reset.
    - Super admin can reset anyone except themselves (use change-password for self).
    - Clinic admin can reset users in their own facility only (not super admins).
    - Generates a temp password if none provided, forces must_change_password=TRUE,
      sends email, and returns the new password so the admin can share it manually
      if email delivery fails."""
    if user["role"] not in [ROLE_SUPER_ADMIN, ROLE_CLINIC_ADMIN]:
        raise HTTPException(403, "Only admins can reset passwords")
    if user["id"] == user_id:
        raise HTTPException(400, "Use Change Password to update your own password")

    target = get_user_by_id(user_id)
    if not target:
        raise HTTPException(404, "User not found")

    # Clinic admins: scope check
    if user["role"] == ROLE_CLINIC_ADMIN:
        if target.get("facility_id") != user.get("facility_id"):
            raise HTTPException(403, "Cannot reset users from other facilities")
        if target.get("role") == ROLE_SUPER_ADMIN:
            raise HTTPException(403, "Cannot reset super admin passwords")

    new_password = (req.new_password or "").strip() or _generate_temp_password()
    if len(new_password) < 8:
        raise HTTPException(400, "Password must be at least 8 characters")

    update_user_password(user_id, new_password, must_change=True)

    email_sent = False
    try:
        email_sent = send_password_reset_email(
            to_email=target["email"],
            full_name=target.get("full_name") or target["email"],
            temp_password=new_password,
            reset_by=user.get("full_name") or user["email"],
        )
    except Exception as e:
        log.error(f"Failed to send password reset email: {e}")

    try:
        log_audit(
            facility_id=target.get("facility_id"),
            action="password_reset",
            details={
                "target_user_id": user_id,
                "target_email": target["email"],
                "email_sent": email_sent,
            },
            user_email=user.get("email"),
            ip_address=_get_client_ip(request),
        )
    except Exception as e:
        log.error(f"Audit log failed for password reset: {e}")

    return {
        "success": True,
        "temp_password": new_password,
        "email_sent": email_sent,
        "must_change_password": True,
    }


# ─────────────────────────────────────────────────────────────────────────
# CLINIC DASHBOARD — auto-load history for logged-in clinic user
# ─────────────────────────────────────────────────────────────────────────

@app.get("/api/clinic/dashboard")
def clinic_dashboard(user: dict = Depends(get_current_user)):
    """Return the logged-in user's clinic data: all quarters of KPI history,
    plus metadata needed to render the dashboard immediately after login.
    Returns empty structure if no uploads exist yet."""
    facility_id = user.get("facility_id")
    facility_name = user.get("facility_name", "")

    if not facility_id:
        # Super admins don't belong to a clinic
        return {"has_data": False, "facility_name": "", "quarters": [], "history": {}, "latest": None}

    history = {}
    if USE_DB:
        try:
            history = get_facility_history_by_id(facility_id)
        except Exception as e:
            log.error(f"Failed to load clinic history: {e}")

    if not history:
        # Fall back to in-memory
        history_key = facility_name.strip().lower()
        history = facility_history.get(history_key, {})

    quarters = sorted(history.keys())
    latest_quarter = quarters[-1] if quarters else None
    latest = None
    if latest_quarter and latest_quarter in history:
        q_data = history[latest_quarter]
        latest = {
            "quarter": latest_quarter,
            "facility": facility_name,
            "kpis": q_data.get("kpis", {}),
            "jawda_summary": q_data.get("jawda_summary", {}),
            "total_records": q_data.get("total_records", 0),
            "col_mapping": q_data.get("col_mapping", {}),
            "data_quality": q_data.get("data_quality", {}),
            "history": history,
        }

    return {
        "has_data": len(quarters) > 0,
        "facility_name": facility_name,
        "quarters": quarters,
        "history": history,
        "latest": latest,
    }


@app.get("/api/notifications")
def list_notifications(user: dict = Depends(get_current_user)):
    """Get notifications for the logged-in user's facility."""
    facility_id = user.get("facility_id")
    if not facility_id:
        return {"notifications": [], "unread": 0}
    # Generate deadline reminders on each check
    if USE_DB:
        try:
            generate_deadline_notifications(facility_id)
        except Exception as e:
            log.error(f"Deadline notification generation failed: {e}")
    try:
        notifications = get_notifications(facility_id, user.get("id"))
        unread = get_unread_count(facility_id, user.get("id"))
        return {"notifications": notifications, "unread": unread}
    except Exception as e:
        log.error(f"Get notifications failed: {e}")
        return {"notifications": [], "unread": 0}


@app.post("/api/notifications/read")
def read_notifications(user: dict = Depends(get_current_user)):
    """Mark all notifications as read."""
    facility_id = user.get("facility_id")
    if not facility_id:
        return {"success": True}
    try:
        mark_notifications_read(facility_id, user.get("id"))
    except Exception as e:
        log.error(f"Mark notifications read failed: {e}")
    return {"success": True}


@app.post("/api/clinic/report/pdf")
def generate_pdf_report(request: Request, quarter: str = Form(...),
                         user: dict = Depends(get_current_user)):
    """Generate a branded PDF report for a quarter."""
    from report_generator import generate_pdf
    from fastapi.responses import Response

    facility_id = user.get("facility_id")
    facility_name = user.get("facility_name", "Clinic")

    if not facility_id:
        raise HTTPException(400, "No facility associated with your account")

    history = {}
    if USE_DB:
        try:
            history = get_facility_history_by_id(facility_id)
        except Exception as e:
            log.error(f"Failed to load history for PDF: {e}")

    if quarter not in history:
        raise HTTPException(404, f"No data found for {quarter}")

    q_data = history[quarter]
    kpis = q_data.get("kpis", {})
    summary = q_data.get("jawda_summary", {})

    try:
        pdf_bytes, report_hash = generate_pdf(
            facility_name=facility_name,
            quarter=quarter,
            kpis=kpis,
            summary=summary,
            history=history,
            generated_by=user.get("full_name") or user.get("email", ""),
        )
    except Exception as e:
        log.error(f"PDF generation failed: {e}")
        raise HTTPException(500, f"PDF generation failed: {str(e)}")

    try:
        log_audit(
            facility_id=facility_id,
            action="report_generated",
            quarter=quarter,
            details={"format": "pdf", "hash": report_hash},
            user_email=user.get("email"),
            ip_address=_get_client_ip(request),
        )
    except Exception as e:
        log.error(f"Audit log for report: {e}")

    filename = f"Jawda-KPI-{facility_name.replace(' ', '-')}-{quarter.replace(' ', '-')}.pdf"
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.post("/api/clinic/submission")
def update_submission(req: UpdateSubmissionRequest, request: Request,
                      user: dict = Depends(get_current_user)):
    """Update submission status for a quarter.
    Clinic admin or quality officer can mark quarters as reviewed/submitted."""
    if user["role"] not in [ROLE_SUPER_ADMIN, ROLE_CLINIC_ADMIN, ROLE_QUALITY_OFFICER]:
        raise HTTPException(403, "Insufficient permissions")
    facility_id = user.get("facility_id")
    if not facility_id:
        raise HTTPException(400, "No facility associated with your account")
    try:
        update_submission_status(
            facility_id=facility_id,
            quarter=req.quarter,
            status=req.status,
            notes=req.notes or "",
            updated_by=user.get("email", ""),
        )
    except ValueError as e:
        raise HTTPException(400, str(e))
    try:
        log_audit(
            facility_id=facility_id,
            action="submission_status_update",
            quarter=req.quarter,
            details={"status": req.status, "notes": req.notes},
            user_email=user.get("email"),
            ip_address=_get_client_ip(request),
        )
    except Exception as e:
        log.error(f"Audit log error: {e}")
    return {"success": True, "quarter": req.quarter, "status": req.status}


@app.post("/api/validate")
async def validate_file(file: UploadFile = File(...)):
    """
    Step 1: Upload file for pre-validation.
    Returns file info, detected columns, quarters, and any issues — without running KPI calculations.
    """

    # File type check
    allowed = [".csv", ".xlsx", ".xls"]
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in allowed:
        raise HTTPException(400, f"Unsupported file type: {ext}. Please upload CSV or Excel (.xlsx, .xls).")

    # File size check
    if file.size and file.size > 50 * 1024 * 1024:
        raise HTTPException(400, "File too large. Maximum 50MB.")

    # Save temp file
    with tempfile.NamedTemporaryFile(delete=False, suffix=ext) as tmp:
        content = await file.read()
        tmp.write(content)
        tmp_path = tmp.name

    try:
        # Load
        df = load_file(tmp_path)

        # Empty check
        if len(df) == 0:
            raise HTTPException(400, "File is empty — no data rows found.")

        if len(df.columns) < 3:
            raise HTTPException(400, f"File has only {len(df.columns)} columns. Expected a patient visit export with multiple fields.")

        # Normalise and map columns
        df = normalise_df(df)
        col_map = map_columns(df)
        mapped = {k: v for k, v in col_map.items() if v is not None}

        if len(mapped) < 2:
            raise HTTPException(400, "Could not recognise enough columns. Make sure the file contains patient visit data with fields like Patient ID, Age, ICD code, BP, etc.")

        # Detect quarters and years
        date_col = col_map.get("date_of_service")
        quarters = []
        years = []
        if date_col:
            quarters = _detect_quarters(df, date_col)
            # Extract unique years
            year_counts = {}
            for val in df[date_col].dropna():
                try:
                    yr = pd.to_datetime(val).year
                    year_counts[yr] = year_counts.get(yr, 0) + 1
                except:
                    continue
            years = [{"year": y, "record_count": c} for y, c in sorted(year_counts.items(), reverse=True)]

        # Warnings
        warnings = []
        if not col_map.get("patient_id"):
            warnings.append("No Patient ID / MRN column detected — patients cannot be deduplicated.")
        if not col_map.get("icd_code"):
            warnings.append("No ICD-10 diagnosis code column detected — most KPIs require this.")
        if not date_col:
            warnings.append("No date column detected — cannot determine quarter automatically.")
        if len(df) < 10:
            warnings.append(f"Only {len(df)} rows found. Results may not be representative.")

        # Store validated file for step 2
        file_id = str(uuid.uuid4())[:8]
        validated_files[file_id] = {
            "path": tmp_path,
            "created": datetime.now().isoformat(),
        }

        return {
            "valid": True,
            "file_id": file_id,
            "filename": file.filename,
            "file_size_kb": round(len(content) / 1024, 1),
            "total_rows": len(df),
            "total_columns": len(df.columns),
            "mapped_fields": len(mapped),
            "mapped_details": mapped,
            "quarters_detected": quarters,
            "years_detected": years,
            "date_range": {
                "earliest": min(pd.to_datetime(df[date_col].dropna())).strftime("%d %b %Y"),
                "latest": max(pd.to_datetime(df[date_col].dropna())).strftime("%d %b %Y"),
            } if date_col and len(df[date_col].dropna()) > 0 else None,
            "warnings": warnings,
        }

    except HTTPException:
        os.unlink(tmp_path)
        raise
    except Exception as e:
        os.unlink(tmp_path)
        raise HTTPException(500, f"Error reading file: {str(e)}")


@app.post("/api/calculate")
async def calculate_kpis(
    file_id: str = Form(...),
    quarter: str = Form(default=""),
    facility_name: str = Form(default="Clinic")
):
    """
    Step 2: Run KPI calculations on a previously validated file.
    If quarter is provided, filters data to that quarter only.
    """

    if file_id not in validated_files:
        raise HTTPException(400, "File not found or expired. Please upload again.")

    tmp_path = validated_files[file_id]["path"]
    if not os.path.exists(tmp_path):
        del validated_files[file_id]
        raise HTTPException(400, "File expired. Please upload again.")

    try:
        df = load_file(tmp_path)
        df_norm = normalise_df(df)
        col_map = map_columns(df_norm)

        # Filter by quarter if specified
        date_col = col_map.get("date_of_service")
        quarter_label = quarter
        if quarter and date_col:
            df_filtered = _filter_by_quarter(df_norm, date_col, quarter)
            if len(df_filtered) == 0:
                raise HTTPException(400, f"No records found for {quarter}. Check the quarter selection.")
            results = run_all_kpis(df_filtered, quarter)
        else:
            # No quarter filter — use all data
            if not quarter_label:
                # Auto-detect most common quarter
                if date_col:
                    quarters = _detect_quarters(df_norm, date_col)
                    quarter_label = quarters[0]["quarter"] if quarters else "Unknown"
                else:
                    quarter_label = "Unknown"
            results = run_all_kpis(df_norm, quarter_label)

        results["facility"] = facility_name
        results["filename"] = os.path.basename(tmp_path)

        run_id = str(uuid.uuid4())[:8]
        results_store[run_id] = results

        _save_to_history(facility_name, results.get("quarter", "Unknown"), results)
        history_key = facility_name.strip().lower()
        if USE_DB:
            try:
                results["history"] = get_facility_history(facility_name)
            except:
                results["history"] = facility_history.get(history_key, {})
        else:
            results["history"] = facility_history.get(history_key, {})

        return {
            "success": True,
            "run_id": run_id,
            "results": results
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"Error processing file: {str(e)}")
    finally:
        # Clean up
        try:
            os.unlink(tmp_path)
        except:
            pass
        validated_files.pop(file_id, None)


def _detect_file_type(df: pd.DataFrame) -> str:
    """Detect file type from column headers."""
    cols_lower = set(str(c).lower().strip().replace("\t","") for c in df.columns)
    # Time Data: has duration columns
    if any("duration from reception" in c for c in cols_lower) or any("duration_from_reception" in c for c in cols_lower):
        return "time_data"
    # E-Claims: has claim-specific financial columns
    if "gross amount" in cols_lower or "gross_amount" in cols_lower or "net amount" in cols_lower:
        return "eclaims"
    # KPI Excel: has BMI + BP READING + OPIOD columns
    has_bmi = any("bmi" in c for c in cols_lower)
    has_bp_reading = any("bp reading" in c or "bp_reading" in c for c in cols_lower)
    if has_bmi and has_bp_reading:
        return "kpi_data"
    # Visit Details: has Primary Diagnosis + E&M/ENM
    has_diag = any("primary diagnosis" in c or "primary_diagnosis" in c for c in cols_lower)
    has_enm = any("e&m" in c or "enm" in c for c in cols_lower)
    if has_diag and has_enm:
        return "visit_details"
    return "unknown"


def _save_temp_file(content: bytes, ext: str) -> str:
    """Save uploaded content to temp file, return path."""
    with tempfile.NamedTemporaryFile(delete=False, suffix=ext) as tmp:
        tmp.write(content)
        return tmp.name


async def _validate_single(file: UploadFile) -> dict:
    """Validate a single file — returns info dict with file_type, path, df stats."""
    allowed = [".csv", ".xlsx", ".xls"]
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in allowed:
        return {"valid": False, "error": f"Unsupported type: {ext}", "filename": file.filename}

    content = await file.read()
    if len(content) > 50 * 1024 * 1024:
        return {"valid": False, "error": "File too large (max 50MB)", "filename": file.filename}

    tmp_path = _save_temp_file(content, ext)
    try:
        df = load_file(tmp_path)
        if len(df) == 0:
            os.unlink(tmp_path)
            return {"valid": False, "error": "File is empty", "filename": file.filename}

        file_type = _detect_file_type(df)
        df_norm = normalise_df(df)
        col_map = map_columns(df_norm)
        mapped = {k: v for k, v in col_map.items() if v is not None}

        return {
            "valid": True,
            "filename": file.filename,
            "file_type": file_type,
            "rows": len(df),
            "columns": len(df.columns),
            "mapped_fields": len(mapped),
            "mapped_details": mapped,
            "path": tmp_path,
        }
    except Exception as e:
        os.unlink(tmp_path)
        return {"valid": False, "error": str(e), "filename": file.filename}


@app.post("/api/validate-multi")
async def validate_multi(
    kpi_data: Optional[UploadFile] = File(None),
    visit_details: Optional[UploadFile] = File(None),
    time_data: Optional[UploadFile] = File(None),
    eclaims: Optional[UploadFile] = File(None),
    prev_kpi_data: Optional[UploadFile] = File(None),
    prev_visit_details: Optional[UploadFile] = File(None),
    prev_time_data: Optional[UploadFile] = File(None),
    prev_eclaims: Optional[UploadFile] = File(None),
):
    """Validate up to 8 files — 4 current quarter + 4 previous quarter for comparison."""
    _cleanup_stale_sessions()

    has_current = kpi_data or visit_details or time_data or eclaims
    has_previous = prev_kpi_data or prev_visit_details or prev_time_data or prev_eclaims
    if not has_current and not has_previous:
        raise HTTPException(400, "Please upload at least one file.")

    results = {}
    file_paths = {}
    prev_results = {}
    prev_file_paths = {}

    # Validate current quarter files
    for slot, file in [("kpi_data", kpi_data), ("visit_details", visit_details),
                        ("time_data", time_data), ("eclaims", eclaims)]:
        if file and file.filename:
            info = await _validate_single(file)
            results[slot] = info
            if info.get("valid") and info.get("path"):
                file_paths[slot] = info["path"]

    # Validate previous quarter files
    for slot, file in [("kpi_data", prev_kpi_data), ("visit_details", prev_visit_details),
                        ("time_data", prev_time_data), ("eclaims", prev_eclaims)]:
        if file and file.filename:
            info = await _validate_single(file)
            prev_results[slot] = info
            if info.get("valid") and info.get("path"):
                prev_file_paths[slot] = info["path"]

    if not any(r.get("valid") for r in results.values()) and not any(r.get("valid") for r in prev_results.values()):
        raise HTTPException(400, "No valid files uploaded.")

    # Detect quarters from whichever file has dates
    quarters = []
    years = []
    date_range = None
    for slot in ["kpi_data", "visit_details", "time_data", "eclaims"]:
        if slot in file_paths:
            try:
                df = load_file(file_paths[slot])
                df_norm = normalise_df(df)
                col_map = map_columns(df_norm)
                date_col = col_map.get("date_of_service")
                if date_col and df_norm[date_col].notna().sum() > 0:
                    quarters = _detect_quarters(df_norm, date_col)
                    year_counts = {}
                    for val in df_norm[date_col].dropna():
                        try:
                            yr = pd.to_datetime(val).year
                            year_counts[yr] = year_counts.get(yr, 0) + 1
                        except:
                            continue
                    years = [{"year": y, "record_count": c} for y, c in sorted(year_counts.items(), reverse=True)]
                    dates = pd.to_datetime(df_norm[date_col].dropna())
                    date_range = {"earliest": dates.min().strftime("%d %b %Y"), "latest": dates.max().strftime("%d %b %Y")}
                    break
            except:
                continue

    # Store all paths for calculate step
    session_id = str(uuid.uuid4())[:8]
    validated_files[session_id] = {
        "paths": file_paths,
        "prev_paths": prev_file_paths,
        "created": datetime.now().isoformat(),
    }

    # Warnings
    warnings = []
    if "kpi_data" not in results or not results.get("kpi_data", {}).get("valid"):
        warnings.append("KPI Data (Excel) not uploaded — BMI, BP, HbA1c, eGFR fields will be missing.")
    if "time_data" not in results or not results.get("time_data", {}).get("valid"):
        warnings.append("Time Data not uploaded — OMC003 (wait time) cannot be calculated.")
    if "visit_details" not in results or not results.get("visit_details", {}).get("valid"):
        warnings.append("Visit Details not uploaded — ICD/CPT cross-validation not available.")

    # Cross-file ID validation: check if patient IDs overlap between files
    id_warnings = []
    if "kpi_data" in file_paths and results.get("kpi_data", {}).get("valid"):
        try:
            df_primary = load_file(file_paths["kpi_data"])
            df_primary = normalise_df(df_primary)
            primary_map = map_columns(df_primary)
            primary_pid = primary_map.get("patient_id")
            primary_fno = primary_map.get("file_no")

            if primary_pid or primary_fno:
                # Build all primary ID sets for cross-checking
                def _id_set(df, col):
                    if col and col in df.columns:
                        return set(df[col].dropna().astype(str).str.strip()
                                   .str.replace(r'\.0$', '', regex=True))
                    return set()

                primary_pid_ids = _id_set(df_primary, primary_pid)
                primary_fno_ids = _id_set(df_primary, primary_fno)

                for slot_name, slot_label, affects in [
                    ("time_data", "Time Data", "OMC003 (wait time)"),
                    ("visit_details", "Visit Details", "ICD/drug enrichment for OMC001/002/005/006"),
                ]:
                    if slot_name in file_paths and results.get(slot_name, {}).get("valid"):
                        try:
                            df_other = load_file(file_paths[slot_name])
                            df_other = normalise_df(df_other)
                            other_map = map_columns(df_other)
                            other_fno_ids = _id_set(df_other, other_map.get("file_no"))
                            other_pid_ids = _id_set(df_other, other_map.get("patient_id"))

                            # Try all 4 combinations, pick the best overlap
                            best_overlap = 0
                            best_total = 1
                            best_desc = ""
                            for p_ids, p_label in [(primary_fno_ids, "file_no"), (primary_pid_ids, "mrn")]:
                                for o_ids, o_label in [(other_fno_ids, "file_no"), (other_pid_ids, "mrn")]:
                                    if p_ids and o_ids:
                                        ov = len(p_ids & o_ids)
                                        if ov > best_overlap:
                                            best_overlap = ov
                                            best_total = len(p_ids)
                                            best_desc = f"KPI Excel {p_label} → {slot_label} {o_label}"

                            overlap_pct = round(best_overlap / best_total * 100) if best_total else 0

                            if overlap_pct < 10:
                                other_sample_col = other_map.get("file_no") or other_map.get("patient_id")
                                other_samples = list(other_fno_ids or other_pid_ids)[:3]
                                primary_samples = list(primary_fno_ids or primary_pid_ids)[:3]
                                id_warnings.append({
                                    "file": slot_label,
                                    "severity": "error",
                                    "message": f"{slot_label} patient IDs do not match KPI Excel ({best_overlap}/{best_total} overlap = {overlap_pct}%). "
                                               f"This will affect {affects}. "
                                               f"KPI Excel samples: {primary_samples}. "
                                               f"{slot_label} samples: {other_samples}. "
                                               f"Ensure the same patient ID format is used across all files."
                                })
                            elif overlap_pct < 50:
                                id_warnings.append({
                                    "file": slot_label,
                                    "severity": "warning",
                                    "message": f"{slot_label} has partial ID overlap with KPI Excel ({overlap_pct}% via {best_desc}). Some patient data may not merge correctly."
                                })
                            # else: good overlap, no warning needed
                        except Exception as e:
                            log.error(f"ID validation for {slot_name}: {e}")
        except Exception as e:
            log.error(f"Cross-file ID validation failed: {e}")

    for w in id_warnings:
        warnings.append(w["message"])

    # Duplicate MRN detection + Column completeness check
    data_quality_warnings = []
    if "kpi_data" in file_paths and results.get("kpi_data", {}).get("valid"):
        try:
            if not 'df_primary' in dir():
                df_primary = load_file(file_paths["kpi_data"])
                df_primary = normalise_df(df_primary)
                primary_map = map_columns(df_primary)

            pid_col = primary_map.get("patient_id")
            age_col = primary_map.get("age")

            # Duplicate MRN with different ages
            if pid_col and age_col:
                import re as _re
                patient_ages = {}
                for _, row in df_primary.iterrows():
                    pid = str(row.get(pid_col, "")).strip()
                    age_val = str(row.get(age_col, ""))
                    m = _re.search(r'(\d+)', age_val)
                    age = int(m.group(1)) if m else None
                    if pid and age is not None:
                        if pid not in patient_ages:
                            patient_ages[pid] = set()
                        patient_ages[pid].add(age)

                suspicious = {pid: ages for pid, ages in patient_ages.items()
                              if len(ages) > 1 and (max(ages) - min(ages)) > 5}
                if suspicious:
                    sample = list(suspicious.items())[:3]
                    sample_str = "; ".join(f"ID {pid}: ages {sorted(ages)}" for pid, ages in sample)
                    data_quality_warnings.append({
                        "type": "duplicate_mrn",
                        "severity": "warning",
                        "count": len(suspicious),
                        "message": f"{len(suspicious)} patient IDs appear with significantly different ages "
                                   f"(may be different patients sharing the same ID). "
                                   f"Examples: {sample_str}. "
                                   f"This affects per-patient deduplication accuracy."
                    })

            # Column completeness — which KPIs can/cannot run
            kpi_column_requirements = {
                "OMC001": {"needs": ["icd_code", "drug_name"], "label": "Asthma Medication Ratio"},
                "OMC003": {"needs": ["wait_minutes"], "label": "Time to See Physician", "alt_file": "Time Data"},
                "OMC004": {"needs": ["bmi"], "label": "BMI Assessment"},
                "OMC005": {"needs": ["hba1c"], "label": "Diabetes HbA1c Control"},
                "OMC006": {"needs": ["bp_reading"], "label": "Controlling High BP"},
                "OMC007": {"needs": ["opioid"], "label": "Opioid Risk"},
                "OMC008": {"needs": ["egfr"], "label": "Kidney Disease Evaluation"},
            }
            missing_kpis = []
            for kpi_id, req in kpi_column_requirements.items():
                has_any = any(primary_map.get(col) for col in req["needs"])
                if not has_any and "alt_file" not in req:
                    missing_kpis.append(f"{kpi_id} ({req['label']})")

            if missing_kpis:
                data_quality_warnings.append({
                    "type": "missing_columns",
                    "severity": "info",
                    "message": f"KPI Excel is missing columns needed for: {', '.join(missing_kpis)}. "
                               f"These KPIs will show as 'Insufficient Data'. "
                               f"Ask your HIS vendor to include these fields in the export."
                })
        except Exception as e:
            log.error(f"Data quality validation failed: {e}")

    for w in data_quality_warnings:
        warnings.append(w["message"])

    total_rows = sum(r.get("rows", 0) for r in results.values() if r.get("valid"))
    prev_total_rows = sum(r.get("rows", 0) for r in prev_results.values() if r.get("valid"))

    # Build column mapping across all uploaded files for the mapping review UI
    from engine.kpi_engine import FIELD_DEFINITIONS
    combined_mapping = {}
    all_available_columns = []
    try:
        for slot, path in file_paths.items():
            df_slot = normalise_df(load_file(path))
            slot_cols = list(df_slot.columns)
            all_available_columns.extend(slot_cols)
            slot_map = map_columns(df_slot)
            for field_key, detected_col in slot_map.items():
                if detected_col and field_key not in combined_mapping:
                    combined_mapping[field_key] = detected_col
        all_available_columns = sorted(set(all_available_columns))
    except Exception as e:
        log.error(f"Column mapping build failed: {e}")

    # Simplify field definitions for frontend
    field_defs_simple = {
        k: {"label": v["label"], "required": v["required"], "file": v["file"], "kpis": v["kpis"]}
        for k, v in FIELD_DEFINITIONS.items()
    }

    return {
        "valid": True,
        "session_id": session_id,
        "files": results,
        "prev_files": prev_results,
        "has_previous": len(prev_file_paths) > 0,
        "total_rows": total_rows,
        "prev_total_rows": prev_total_rows,
        "quarters_detected": quarters,
        "years_detected": years,
        "date_range": date_range,
        "warnings": warnings,
        "id_warnings": id_warnings,
        "data_quality_warnings": data_quality_warnings,
        "col_mapping": combined_mapping,
        "available_columns": all_available_columns,
        "field_definitions": field_defs_simple,
    }


@app.post("/api/calculate-multi")
async def calculate_multi(
    request: Request,
    session_id: str = Form(...),
    quarter: str = Form(default=""),
    facility_name: str = Form(default="Clinic"),
    col_mapping: Optional[str] = Form(default=None),
    user: dict = Depends(get_current_user),
):
    """Run KPI calculations on previously validated multi-file session.
    Merges files by patient ID where possible.
    col_mapping: optional JSON string of user-corrected column mapping."""

    # Use facility from logged-in user when available (multi-tenancy)
    if user.get("facility_name"):
        facility_name = user["facility_name"]

    client_ip = _get_client_ip(request)
    user_agent = request.headers.get("user-agent", "")

    # Parse user-provided column mapping override
    user_col_mapping = None
    if col_mapping:
        try:
            user_col_mapping = json.loads(col_mapping)
            log.info(f"Using user-provided column mapping: {len(user_col_mapping)} fields")
        except Exception as e:
            log.warning(f"Invalid col_mapping JSON: {e}")

    if session_id not in validated_files:
        raise HTTPException(400, "Session not found or expired. Please upload again.")

    session = validated_files[session_id]
    paths = session.get("paths", {})

    if not paths:
        raise HTTPException(400, "No files in session.")

    try:
        # Load the primary KPI data file first
        primary_slot = "kpi_data" if "kpi_data" in paths else list(paths.keys())[0]
        df_primary = load_file(paths[primary_slot])
        df_merged = normalise_df(df_primary)
        col_map = map_columns(df_merged)
        pid_col = col_map.get("patient_id")
        log.info(f"Primary ({primary_slot}): {len(df_merged)} rows, columns={list(df_merged.columns)[:20]}")
        log.info(f"Primary mapped: patient_id={pid_col}, file_no={col_map.get('file_no')}")
        merge_diagnostics = []

        def _normalize_join_key(series):
            """Normalize join keys: strip, lowercase, remove leading zeros, remove .0 suffix."""
            return (series.astype(str).str.strip()
                    .str.replace(r'\.0$', '', regex=True)  # 156.0 → 156
                    .str.lower()
                    .str.lstrip('0')
                    .replace('', 'EMPTY'))

        def _try_join(df_left, df_right, left_col, right_col, value_cols, label):
            """Try to join and return (merged_df, match_count, total).
            Tries exact match first, then normalized match.
            Never drops existing columns from df_left."""
            total = len(df_left)

            # Only bring value columns that don't already exist in df_left
            new_value_cols = [c for c in value_cols if c not in df_left.columns]
            if not new_value_cols:
                return df_left, 0, total

            # Strategy 1: exact string match
            right_sub = df_right[[right_col] + new_value_cols].copy()
            right_sub[right_col] = right_sub[right_col].astype(str).str.strip()
            df_left = df_left.copy()
            df_left[left_col] = df_left[left_col].astype(str).str.strip()
            right_sub = right_sub.rename(columns={right_col: left_col})
            right_sub = right_sub.drop_duplicates(subset=[left_col], keep='first')
            merged = df_left.merge(right_sub, on=left_col, how='left', suffixes=('', f'_{label}'))
            matched = sum(merged[new_value_cols[0]].notna()) if new_value_cols[0] in merged.columns else 0

            if matched > total * 0.3:  # >30% match = good enough
                return merged, matched, total

            # Strategy 2: normalized keys (strip zeros, lowercase)
            log.info(f"  {label}: exact match {matched}/{total}, trying normalized keys...")
            # Drop only the NEW columns that Strategy 1 added (not existing ones)
            df_left_clean = merged.drop(columns=[c for c in new_value_cols if c in merged.columns], errors='ignore')
            right_sub2 = df_right[[right_col] + new_value_cols].copy()
            join_key = f'_norm_join_{label}'
            df_left_clean[join_key] = _normalize_join_key(df_left_clean[left_col])
            right_sub2[join_key] = _normalize_join_key(right_sub2[right_col])
            right_sub2 = right_sub2.drop(columns=[right_col])
            right_sub2 = right_sub2.drop_duplicates(subset=[join_key], keep='first')
            merged2 = df_left_clean.merge(right_sub2, on=join_key, how='left', suffixes=('', f'_{label}'))
            matched2 = sum(merged2[new_value_cols[0]].notna()) if new_value_cols[0] in merged2.columns else 0
            merged2 = merged2.drop(columns=[join_key], errors='ignore')

            if matched2 > matched:
                log.info(f"  {label}: normalized match improved {matched} → {matched2}/{total}")
                return merged2, matched2, total

            return merged, matched, total

        # ── Merge Time Data ─────────────────────────────────────────────
        time_merge_failed = False
        df_time_standalone = None  # Always kept for OMC003 — Time Data is the source of truth for wait times
        if "time_data" in paths and "time_data" != primary_slot:
            df_time = load_file(paths["time_data"])
            df_time_norm = normalise_df(df_time)
            time_col_map = map_columns(df_time_norm)
            log.info(f"  Time Data: {len(df_time_norm)} rows, columns={list(df_time_norm.columns)}")
            log.info(f"  Time Data mapped: { {k:v for k,v in time_col_map.items() if v} }")
            wait_col = time_col_map.get("wait_minutes")
            reg_col = time_col_map.get("registration_ts")
            consult_col = time_col_map.get("consult_ts")

            # Collect all useful time columns
            time_value_cols = [c for c in [wait_col, reg_col, consult_col] if c]

            # Always keep Time Data for standalone OMC003 calculation
            if time_value_cols:
                df_time_standalone = df_time_norm
                # Try multiple join key strategies — always require a real key match
                time_file_col = time_col_map.get("file_no") or time_col_map.get("patient_id")
                primary_file_col = col_map.get("file_no") or col_map.get("patient_id")

                # Debug: log sample IDs from both sides to diagnose mismatches
                if time_file_col and primary_file_col:
                    try:
                        time_ids = df_time_norm[time_file_col].dropna().astype(str).str.strip().head(5).tolist()
                        primary_ids = df_merged[primary_file_col].dropna().astype(str).str.strip().head(5).tolist()
                        log.info(f"  Time Data join debug: time_col='{time_file_col}' samples={time_ids}, "
                                 f"primary_col='{primary_file_col}' samples={primary_ids}")
                    except Exception as e:
                        log.info(f"  Time Data join debug error: {e}")

                matched = 0
                if time_file_col and primary_file_col:
                    # Strategy 1+2: exact then normalized join on file_no→file_no
                    df_merged, matched, total = _try_join(
                        df_merged, df_time_norm, primary_file_col, time_file_col, time_value_cols, 'time')
                    merge_diagnostics.append(f"Time Data: {matched}/{total} rows matched on {primary_file_col}")

                    # Strategy 3: try time's file_no against primary's patient_id (MRN)
                    if matched < len(df_merged) * 0.1 and pid_col and pid_col != primary_file_col:
                        log.info(f"  Time Data: trying time file_no → primary patient_id ({pid_col})")
                        for vc in time_value_cols:
                            df_merged = df_merged.drop(columns=[vc], errors='ignore')
                        df_merged, matched2, total = _try_join(
                            df_merged, df_time_norm, pid_col, time_file_col, time_value_cols, 'time')
                        if matched2 > matched:
                            matched = matched2
                            merge_diagnostics[-1] = f"Time Data: {matched}/{total} rows matched on {pid_col} (time file_no→primary MRN)"

                    # Strategy 4: try time's patient_id against primary's file_no
                    if matched < len(df_merged) * 0.1:
                        alt_time_col = time_col_map.get("patient_id")
                        if alt_time_col and alt_time_col != time_file_col:
                            log.info(f"  Time Data: trying time patient_id → primary file_no ({primary_file_col})")
                            for vc in time_value_cols:
                                df_merged = df_merged.drop(columns=[vc], errors='ignore')
                            df_merged, matched3, total = _try_join(
                                df_merged, df_time_norm, primary_file_col, alt_time_col, time_value_cols, 'time')
                            if matched3 > matched:
                                matched = matched3
                                merge_diagnostics[-1] = f"Time Data: {matched}/{total} rows matched on {primary_file_col} (time MRN→primary file_no)"

                    # Strategy 5: try time's patient_id against primary's patient_id
                    if matched < len(df_merged) * 0.1:
                        alt_time_col = time_col_map.get("patient_id")
                        if alt_time_col and alt_time_col != time_file_col and pid_col:
                            log.info(f"  Time Data: trying time patient_id → primary patient_id ({pid_col})")
                            for vc in time_value_cols:
                                df_merged = df_merged.drop(columns=[vc], errors='ignore')
                            df_merged, matched4, total = _try_join(
                                df_merged, df_time_norm, pid_col, alt_time_col, time_value_cols, 'time')
                            if matched4 > matched:
                                matched = matched4
                                merge_diagnostics[-1] = f"Time Data: {matched}/{total} rows matched on {pid_col} (MRN→MRN)"

                    if matched < len(df_merged) * 0.1:
                        merge_diagnostics[-1] += " — WARNING: very low match rate, check ID formats"
                        time_merge_failed = True
                        df_time_standalone = df_time_norm
                else:
                    merge_diagnostics.append("Time Data: no join key found, skipped")
                    time_merge_failed = True
                    df_time_standalone = df_time_norm
            else:
                merge_diagnostics.append("Time Data: no wait time or timestamp columns found")

        # ── Merge Visit Details ─────────────────────────────────────────
        if "visit_details" in paths and "visit_details" != primary_slot:
            df_visit = load_file(paths["visit_details"])
            df_visit_norm = normalise_df(df_visit)
            visit_col_map = map_columns(df_visit_norm)
            visit_pid = visit_col_map.get("patient_id")

            if visit_pid and pid_col:
                try:
                    visit_ids = df_visit_norm[visit_pid].dropna().astype(str).str.strip().head(5).tolist()
                    primary_pids = df_merged[pid_col].dropna().astype(str).str.strip().head(5).tolist()
                    log.info(f"  Visit Details join debug: visit_col='{visit_pid}' samples={visit_ids}, "
                             f"primary_col='{pid_col}' samples={primary_pids}")
                except:
                    pass
                cols_to_bring = []
                for field_key in ["icd_code", "icd_secondary", "diabetes_flag", "htn_flag",
                                  "pregnancy", "bp_reading", "bmi", "hba1c", "drug_name",
                                  "drug_class", "days_supplied", "egfr", "uacr"]:
                    col = visit_col_map.get(field_key)
                    if col and col != visit_pid and col not in cols_to_bring:
                        # Don't overwrite columns that already exist in primary
                        if col not in df_merged.columns:
                            cols_to_bring.append(col)

                if cols_to_bring:
                    df_merged, matched, total = _try_join(
                        df_merged, df_visit_norm, pid_col, visit_pid, cols_to_bring, 'visit')
                    merge_diagnostics.append(f"Visit Details: {matched}/{total} rows matched, {len(cols_to_bring)} columns added")
                else:
                    merge_diagnostics.append("Visit Details: no new columns to add (all already in primary)")
            else:
                merge_diagnostics.append("Visit Details: no patient ID column found for join")

        # ── Merge E-Claims (if it has useful columns not already present) ──
        if "eclaims" in paths and "eclaims" != primary_slot:
            df_eclaims = load_file(paths["eclaims"])
            df_eclaims_norm = normalise_df(df_eclaims)
            eclaims_col_map = map_columns(df_eclaims_norm)
            eclaims_pid = eclaims_col_map.get("patient_id")

            if eclaims_pid and pid_col:
                eclaims_cols = []
                for field_key in ["age", "date_of_birth", "gender", "visit_type"]:
                    col = eclaims_col_map.get(field_key)
                    if col and col != eclaims_pid and col not in eclaims_cols:
                        if col not in df_merged.columns:
                            eclaims_cols.append(col)

                if eclaims_cols:
                    df_merged, matched, total = _try_join(
                        df_merged, df_eclaims_norm, pid_col, eclaims_pid, eclaims_cols, 'eclaims')
                    merge_diagnostics.append(f"E-Claims: {matched}/{total} rows matched, {len(eclaims_cols)} columns added")

        # Re-map columns after merge (new columns may be available now)
        col_map = map_columns(df_merged)

        # ── Post-merge validation ───────────────────────────────────────
        # Check critical columns — warn if they exist but are all NaN
        critical_checks = [
            ("wait_minutes", "OMC003 (wait time)"),
            ("hba1c", "OMC005 (diabetes HbA1c)"),
            ("egfr", "OMC008 (kidney eGFR)"),
        ]
        for field, kpi_label in critical_checks:
            col = col_map.get(field)
            if col and col in df_merged.columns:
                non_null = df_merged[col].notna().sum()
                if non_null == 0:
                    merge_diagnostics.append(f"WARNING: {field} column exists but has no data after merge — {kpi_label} will show 0/0")
                    log.warning(f"Post-merge: {field} column all NaN — {kpi_label} affected")

        # Memory check after merge
        import sys
        merged_mb = df_merged.memory_usage(deep=True).sum() / (1024 * 1024)
        log.info(f"Merged dataframe: {len(df_merged)} rows x {len(df_merged.columns)} cols, {merged_mb:.1f} MB in memory")
        if merged_mb > 500:
            log.warning(f"Large merged dataframe: {merged_mb:.0f} MB — may cause memory pressure")
            merge_diagnostics.append(f"WARNING: Large dataset ({merged_mb:.0f} MB in memory) — processing may be slow")

        log.info(f"Merge diagnostics: {merge_diagnostics}")

        # Filter by quarter if specified
        date_col = col_map.get("date_of_service")
        quarter_label = quarter
        if quarter and date_col:
            df_filtered = _filter_by_quarter(df_merged, date_col, quarter)
            if len(df_filtered) == 0:
                raise HTTPException(400, f"No records found for {quarter}.")
            results = run_all_kpis(df_filtered, quarter, col_mapping_override=user_col_mapping)
        else:
            if not quarter_label and date_col:
                quarters = _detect_quarters(df_merged, date_col)
                quarter_label = quarters[0]["quarter"] if quarters else "Unknown"
            elif not quarter_label:
                quarter_label = "Unknown"
            results = run_all_kpis(df_merged, quarter_label, col_mapping_override=user_col_mapping)

        # ── OMC003: ALWAYS run on Time Data independently ──
        # Time Data is the source of truth for wait times. Each row = 1 visit.
        # Running on the merged dataframe inflates counts due to duplicate
        # patient rows in KPI Excel (multiple visits per patient).
        if df_time_standalone is not None:
            log.info("OMC003: running on Time Data independently (source of truth)")
            try:
                from engine.kpi_engine import calc_omc003 as _calc_omc003, map_columns as _map_cols
                time_cm = _map_cols(df_time_standalone)
                time_date_col = time_cm.get("date_of_service")
                df_time_for_calc = df_time_standalone
                if quarter and time_date_col:
                    df_time_filtered = _filter_by_quarter(df_time_standalone, time_date_col, quarter)
                    if len(df_time_filtered) > 0:
                        df_time_for_calc = df_time_filtered
                omc003_result = _calc_omc003(df_time_for_calc, time_cm)
                omc003_dict = omc003_result.to_dict()
                if omc003_dict.get("denominator", 0) > 0:
                    results["kpis"]["OMC003"] = omc003_dict
                    merge_diagnostics.append(
                        f"OMC003 calculated from Time Data ({len(df_time_for_calc)} visits): "
                        f"{omc003_dict['numerator']}/{omc003_dict['denominator']} "
                        f"({omc003_dict['percentage']}%)")
                    log.info(f"OMC003 from Time Data: {omc003_dict['numerator']}/{omc003_dict['denominator']}")
            except Exception as e:
                log.error(f"OMC003 standalone failed: {e}")
                merge_diagnostics.append(f"OMC003 standalone calculation failed: {str(e)[:100]}")

        # Recalculate jawda_summary (OMC003 may have changed)
        kpi_vals = [v for k, v in results["kpis"].items() if not k.startswith("ERROR")]
        na_count = sum(1 for k in kpi_vals if k["status"] == "not_applicable")
        insuff = sum(1 for k in kpi_vals if k["status"] == "insufficient_data")
        calculable = [k for k in kpi_vals if k["status"] not in ("insufficient_data", "not_applicable")]
        meeting = [k for k in calculable if k.get("meets_target") is True]
        below = [k for k in calculable if k.get("meets_target") is False]
        proxy = [k for k in kpi_vals if k["status"] == "proxy"]
        results["jawda_summary"] = {
            "total_kpis": len(kpi_vals),
            "calculable": len(calculable),
            "meeting_target": len(meeting),
            "below_target": len(below),
            "missing_data": insuff,
            "not_applicable": na_count,
            "proxy_data": len(proxy),
            "readiness_pct": round(len(meeting) / len(calculable) * 100) if calculable else 0,
            "verdict": "ready" if len(calculable) > 0 and len(below) == 0 and insuff == 0
                       else "attention" if len(meeting) > 0
                       else "not_ready",
        }

        results["facility"] = facility_name
        results["files_used"] = list(paths.keys())
        results["merge_diagnostics"] = merge_diagnostics

        # Generate smart recommendations
        try:
            from recommendations import generate_recommendations, generate_predictions
            results["recommendations"] = generate_recommendations(results.get("kpis", {}), results.get("jawda_summary", {}))
            # Predictions need history
            if results.get("history"):
                results["predictions"] = generate_predictions(results["history"])
            else:
                results["predictions"] = []
        except Exception as e:
            log.error(f"Recommendations/predictions failed: {e}")
            results["recommendations"] = []
            results["predictions"] = []

        # If user provided custom column mapping, force-save as clinic default
        if user_col_mapping and user.get("facility_id") and USE_DB:
            try:
                save_facility_col_mapping(user["facility_id"], user_col_mapping, force=True)
                log.info("Saved user-confirmed column mapping as clinic default")
            except Exception as e:
                log.error(f"Failed to save user column mapping: {e}")

        # Save current quarter to history
        _save_to_history(facility_name, results.get("quarter", "Unknown"), results)

        # Send email notification to the user who triggered the calculation
        try:
            if user.get("email") and user["email"] != "anonymous":
                send_kpi_calculation_email(
                    to_email=user["email"],
                    full_name=user.get("full_name") or user["email"],
                    facility_name=facility_name,
                    quarter=results.get("quarter", "Unknown"),
                    summary=results.get("jawda_summary", {}),
                )
        except Exception as e:
            log.error(f"Failed to send KPI calculation email: {e}")

        # Audit log
        if USE_DB:
            try:
                summary = results.get("jawda_summary", {})
                log_audit(
                    facility_name=facility_name,
                    action="kpi_calculation",
                    quarter=results.get("quarter", "Unknown"),
                    user_email=user.get("email"),
                    ip_address=client_ip,
                    user_agent=user_agent,
                    details={
                        "total_records": results.get("total_records", 0),
                        "files_used": list(paths.keys()),
                        "meeting_target": summary.get("meeting_target", 0),
                        "calculable": summary.get("calculable", 0),
                        "verdict": summary.get("verdict", ""),
                        "kpi_results": {
                            kpi_id: {
                                "numerator": kpi.get("numerator", 0),
                                "denominator": kpi.get("denominator", 0),
                                "percentage": kpi.get("percentage", 0),
                                "status": kpi.get("status", ""),
                                "meets_target": kpi.get("meets_target"),
                            }
                            for kpi_id, kpi in results.get("kpis", {}).items()
                            if not kpi_id.startswith("ERROR")
                        }
                    }
                )
            except Exception as e:
                log.error(f"Audit log error: {e}")

        # Process previous quarter files if provided
        prev_paths = session.get("prev_paths", {})
        if prev_paths:
            try:
                prev_primary_slot = "kpi_data" if "kpi_data" in prev_paths else list(prev_paths.keys())[0]
                df_prev = load_file(prev_paths[prev_primary_slot])
                df_prev_merged = normalise_df(df_prev)
                prev_col_map = map_columns(df_prev_merged)

                # Try to merge time data — first by row count, then by ID join
                time_merged = False
                if "time_data" in prev_paths and "time_data" != prev_primary_slot:
                    df_prev_time = load_file(prev_paths["time_data"])
                    df_prev_time_norm = normalise_df(df_prev_time)
                    prev_time_map = map_columns(df_prev_time_norm)
                    prev_wait = prev_time_map.get("wait_minutes")
                    if prev_wait:
                        if len(df_prev_time_norm) == len(df_prev_merged):
                            df_prev_merged[prev_wait] = df_prev_time_norm[prev_wait].values
                            time_merged = True
                        else:
                            # Try join by FILE NO
                            time_file_col = prev_time_map.get("file_no") or prev_time_map.get("patient_id")
                            primary_file_col = prev_col_map.get("file_no")
                            if time_file_col and primary_file_col:
                                time_subset = df_prev_time_norm[[time_file_col, prev_wait]].copy()
                                time_subset[time_file_col] = time_subset[time_file_col].astype(str)
                                df_prev_merged_copy = df_prev_merged.copy()
                                df_prev_merged_copy[primary_file_col] = df_prev_merged_copy[primary_file_col].astype(str)
                                time_subset = time_subset.rename(columns={time_file_col: primary_file_col})
                                time_subset = time_subset.drop_duplicates(subset=[primary_file_col], keep='first')
                                merged = df_prev_merged_copy.merge(time_subset, on=primary_file_col, how='left', suffixes=('', '_time'))
                                # Check if any wait times actually merged
                                if merged[prev_wait].notna().sum() > 0:
                                    df_prev_merged = merged
                                    time_merged = True

                # Merge prev visit details if available — bring ICD, flags, BP, BMI
                if "visit_details" in prev_paths and "visit_details" != prev_primary_slot:
                    df_prev_visit = load_file(prev_paths["visit_details"])
                    df_prev_visit_norm = normalise_df(df_prev_visit)
                    prev_visit_map = map_columns(df_prev_visit_norm)
                    prev_visit_pid = prev_visit_map.get("patient_id")
                    prev_pid_col = map_columns(df_prev_merged).get("patient_id")

                    if prev_visit_pid and prev_pid_col:
                        # Collect all useful columns from visit details
                        cols_to_bring = [prev_visit_pid]
                        for field_key in ["icd_code", "icd_secondary", "diabetes_flag", "htn_flag",
                                          "pregnancy", "bp_reading", "bmi", "hba1c", "drug_name"]:
                            col = prev_visit_map.get(field_key)
                            if col and col != prev_visit_pid and col not in cols_to_bring:
                                cols_to_bring.append(col)

                        if len(cols_to_bring) > 1:
                            pv_subset = df_prev_visit_norm[cols_to_bring].copy()
                            pv_subset[prev_visit_pid] = pv_subset[prev_visit_pid].astype(str)
                            df_prev_merged[prev_pid_col] = df_prev_merged[prev_pid_col].astype(str)
                            pv_subset = pv_subset.rename(columns={prev_visit_pid: prev_pid_col})
                            pv_subset = pv_subset.drop_duplicates(subset=[prev_pid_col], keep='first')
                            df_prev_merged = df_prev_merged.merge(pv_subset, on=prev_pid_col, how='left', suffixes=('', '_visit'))

                prev_col_map = map_columns(df_prev_merged)
                prev_date_col = prev_col_map.get("date_of_service")
                prev_quarter_label = ""
                if prev_date_col:
                    prev_quarters = _detect_quarters(df_prev_merged, prev_date_col)
                    prev_quarter_label = prev_quarters[0]["quarter"] if prev_quarters else "Previous"
                else:
                    prev_quarter_label = "Previous"

                # Run KPIs on merged previous data
                prev_results = run_all_kpis(df_prev_merged, prev_quarter_label)

                # If time data didn't merge (different IDs), run it separately for OMC003
                if not time_merged and "time_data" in prev_paths:
                    df_prev_time = load_file(prev_paths["time_data"])
                    time_only_results = run_all_kpis(df_prev_time, prev_quarter_label)
                    omc003_time = time_only_results.get("kpis", {}).get("OMC003", {})
                    if omc003_time.get("denominator", 0) > 0:
                        prev_results["kpis"]["OMC003"] = omc003_time

                _save_to_history(facility_name, prev_quarter_label, prev_results)
            except Exception as e:
                # Previous quarter processing is best-effort — log and surface but don't block
                log.error(f"Previous quarter processing failed: {e}")
                if "warnings" not in results:
                    results["warnings"] = []
                results["warnings"].append(f"Previous quarter comparison could not be processed: {str(e)[:200]}")

        # Attach full history for frontend
        history_key = facility_name.strip().lower()
        if USE_DB:
            try:
                results["history"] = get_facility_history(facility_name)
            except:
                results["history"] = facility_history.get(history_key, {})
        else:
            results["history"] = facility_history.get(history_key, {})

        run_id = str(uuid.uuid4())[:8]
        results_store[run_id] = results

        return {"success": True, "run_id": run_id, "results": results}

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"Error processing files: {str(e)}")
    finally:
        all_paths = list(paths.values()) + list(session.get("prev_paths", {}).values())
        for p in all_paths:
            try:
                os.unlink(p)
            except:
                pass
        validated_files.pop(session_id, None)


# Keep legacy upload endpoint for backwards compatibility
@app.post("/api/upload")
async def upload_file(
    file: UploadFile = File(...),
    quarter: str = Form(default="Q1 2025"),
    facility_name: str = Form(default="Clinic")
):
    """Legacy single-step upload. Validates and calculates in one call."""

    allowed = [".csv", ".xlsx", ".xls"]
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in allowed:
        raise HTTPException(400, f"File type not supported. Upload CSV or Excel. Got: {ext}")

    if file.size and file.size > 50 * 1024 * 1024:
        raise HTTPException(400, "File too large. Maximum 50MB.")

    with tempfile.NamedTemporaryFile(delete=False, suffix=ext) as tmp:
        content = await file.read()
        tmp.write(content)
        tmp_path = tmp.name

    try:
        df = load_file(tmp_path)
        results = run_all_kpis(df, quarter)
        results["facility"] = facility_name
        results["filename"] = file.filename

        run_id = str(uuid.uuid4())[:8]
        results_store[run_id] = results

        return {
            "success": True,
            "run_id": run_id,
            "results": results
        }

    except Exception as e:
        raise HTTPException(500, f"Error processing file: {str(e)}")
    finally:
        os.unlink(tmp_path)


@app.get("/api/results/{run_id}")
def get_results(run_id: str):
    if run_id not in results_store:
        raise HTTPException(404, "Results not found. They may have expired.")
    return results_store[run_id]

@app.get("/api/history/{facility_name}")
def get_history(facility_name: str):
    """Get all stored quarterly results for a facility."""
    # Try database first, fall back to in-memory
    history = {}
    if USE_DB:
        try:
            history = get_facility_history(facility_name)
        except Exception as e:
            log.error(f"DB history error: {e}")
    if not history:
        key = facility_name.strip().lower()
        history = facility_history.get(key, {})
    if not history:
        raise HTTPException(404, "No history found for this facility.")
    return {
        "facility": facility_name,
        "quarters": history,
        "quarter_list": sorted(history.keys()),
    }

@app.get("/api/audit/{facility_name}")
def get_audit(facility_name: str, limit: int = 50):
    """Get audit trail for a facility — shows all uploads, calculations, and actions."""
    if USE_DB:
        try:
            entries = get_audit_log(facility_name, limit)
            if not entries:
                raise HTTPException(404, "No audit history found for this facility.")
            return {"facility": facility_name, "entries": entries, "total": len(entries)}
        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(500, f"Error retrieving audit log: {str(e)}")
    else:
        raise HTTPException(404, "Audit log requires database connection.")


@app.get("/api/kpi-definitions")
def kpi_definitions():
    """Return official KPI definitions for reference."""
    return {
        "source": "DOH Jawda Guidance v1.4 — Q1 2025",
        "kpis": [
            {"id":"OMC001","title":"Asthma Medication Ratio","domain":"Effectiveness","direction":"Higher","description":"% of asthma patients (5-64) with controller:total medication ratio >= 0.50"},
            {"id":"OMC002","title":"Avoidance of Antibiotics for Bronchitis","domain":"Safety","direction":"Higher","description":"% of bronchitis patients NOT prescribed antibiotics"},
            {"id":"OMC003","title":"Time to See Physician","domain":"Timeliness","direction":"Higher","description":"% seen within 60 min of registration — auto-collected by Malaffi"},
            {"id":"OMC004","title":"Weight/BMI Assessment & Counselling","domain":"Patient-Centredness","direction":"Higher","description":"% of BMI>=25 patients with documented management plan every 6 months"},
            {"id":"OMC005","title":"Diabetes HbA1c Good Control","domain":"Effectiveness","direction":"Higher","description":"% of diabetics (18-75) with HbA1c <= 7.0% in last 12 months"},
            {"id":"OMC006","title":"Controlling High Blood Pressure","domain":"Effectiveness","direction":"Higher","description":"% of hypertension patients with BP < 130/80"},
            {"id":"OMC007","title":"Risk of Continued Opioid Use","domain":"Coordination","direction":"Lower","description":"% of new opioid patients with >= 15 days in 30-day period"},
            {"id":"OMC008","title":"Kidney Disease Evaluation","domain":"Effectiveness","direction":"Higher","description":"% of eGFR<90 patients with both eGFR and uACR tested every 6 months"},
        ]
    }

# Serve frontend static files — must be AFTER all API routes
STATIC_DIR = os.path.join(os.path.dirname(__file__), "static")
if os.path.exists(STATIC_DIR):
    app.mount("/assets", StaticFiles(directory=os.path.join(STATIC_DIR, "assets")), name="assets")

    @app.get("/{full_path:path}")
    async def serve_frontend(full_path: str):
        """Serve the React frontend for all non-API routes."""
        file_path = os.path.join(STATIC_DIR, full_path)
        if os.path.isfile(file_path):
            return FileResponse(file_path)
        return FileResponse(os.path.join(STATIC_DIR, "index.html"))


if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)
